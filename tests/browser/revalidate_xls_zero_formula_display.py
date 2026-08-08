"""Browser regression for BIFF8 formula-zero display and XLSX export.

Requirements:
    pip install playwright beautifulsoup4 openpyxl
    A Chromium executable available as CHROMIUM_PATH or /usr/bin/chromium.

Run from the repository root:
    python3 tests/browser/revalidate_xls_zero_formula_display.py
"""
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from browser_support import launch_browser, requested_browser_name
import json
import os

ROOT = Path(__file__).resolve().parents[2]
FIXTURE = ROOT / "tests" / "compatibility-fixtures" / "spreadsheets" / "independent_biff8_zero_formula_display.xls"
OUT = ROOT / "tests" / "browser" / "results"
OUT.mkdir(parents=True, exist_ok=True)
CHROMIUM = os.environ.get("CHROMIUM_PATH", "/usr/bin/chromium")


def load_app(page):
    soup = BeautifulSoup((ROOT / "apps/spreadsheets/index.html").read_text(), "html.parser")
    for node in soup.find_all(["script", "link"]):
        node.decompose()
    base = soup.new_tag("base", href=ROOT.as_uri() + "/")
    if soup.head:
        soup.head.insert(0, base)
    page.set_content(str(soup), wait_until="domcontentloaded")
    for css in (ROOT / "apps/spreadsheets/styles.css", ROOT / "shared/office-shell.css"):
        page.add_style_tag(path=str(css))
    for js in (
        ROOT / "shared/office-runtime.js",
        ROOT / "shared/vendor/jszip.min.js",
        ROOT / "apps/spreadsheets/xls-biff8-engine.js",
        ROOT / "apps/spreadsheets/xlsx-engine.js",
        ROOT / "shared/office-shell.js",
        ROOT / "apps/spreadsheets/formula-safety.js",
        ROOT / "apps/spreadsheets/history-controller.js",
        ROOT / "apps/spreadsheets/app.js",
    ):
        page.add_script_tag(path=str(js))


def wait_open(page):
    page.wait_for_function(
        "!document.querySelector('#gridViewport').hidden && !document.querySelector('#loading').offsetParent",
        timeout=30000,
    )


def cell_text(page, row, col):
    return page.locator(f'.cell[data-r="{row}"][data-c="{col}"]').inner_text()


def assert_grid_values(page):
    actual = {
        "formula_visible_zero": cell_text(page, 1, 1),
        "formula_hidden_zero": cell_text(page, 2, 1),
        "literal_visible_zero": cell_text(page, 3, 1),
        "literal_hidden_zero": cell_text(page, 4, 1),
        "formula_one": cell_text(page, 5, 1),
    }
    expected = {
        "formula_visible_zero": "0.00",
        "formula_hidden_zero": "",
        "literal_visible_zero": "0.00",
        "literal_hidden_zero": "",
        "formula_one": "1.00",
    }
    if actual != expected:
        raise RuntimeError(f"Unexpected displayed values: {actual!r}")
    return actual


def assert_reopened_values(page):
    actual = {
        "B2": cell_text(page, 1, 1),
        "B3": cell_text(page, 2, 1),
        "B4": cell_text(page, 3, 1),
        "B5": cell_text(page, 4, 1),
        "B6": cell_text(page, 5, 1),
    }
    numeric = {key: float(value) for key, value in actual.items()}
    expected = {"B2": 0.0, "B3": 0.0, "B4": 0.0, "B5": 0.0, "B6": 1.0}
    if numeric != expected:
        raise RuntimeError(f"Reopened values were not preserved: {actual!r}")
    return actual


def main():
    output = OUT / "independent_biff8_zero_formula_display.saved.xlsx"
    errors, dialogs = [], []
    with sync_playwright() as playwright:
        browser = launch_browser(playwright)
        page = browser.new_page(viewport={"width": 1400, "height": 900}, accept_downloads=True)
        page.on("pageerror", lambda error: errors.append(str(error)))
        page.on("dialog", lambda dialog: (dialogs.append(dialog.message), dialog.accept()))
        load_app(page)
        page.set_input_files("#fileInput", str(FIXTURE))
        wait_open(page)
        imported = assert_grid_values(page)
        page.click("#formMode")
        page.wait_for_function("document.body.classList.contains('page-view')")
        page_values = assert_grid_values(page)
        page.click("#gridMode")

        page.click("#saveBtn")
        page.wait_for_function("document.querySelector('#savePanel').style.display==='flex'", timeout=30000)
        with page.expect_download(timeout=30000) as download_info:
            page.click("#downloadBtn")
        download_info.value.save_as(str(output))
        page.close()

        reopened = browser.new_page(viewport={"width": 1400, "height": 900})
        reopened.on("pageerror", lambda error: errors.append(str(error)))
        load_app(reopened)
        reopened.set_input_files("#fileInput", str(output))
        wait_open(reopened)
        reopened_values = assert_reopened_values(reopened)
        reopened.close()
        browser.close()

    if errors or dialogs:
        raise RuntimeError(f"Browser errors: dialogs={dialogs}, errors={errors}")

    wb = load_workbook(output, data_only=False)
    ws = wb["Zero display"]
    workbook_values = {ref: ws[ref].value for ref in ("B2", "B3", "B4", "B5", "B6")}
    workbook_formats = {ref: ws[ref].number_format for ref in ("B2", "B3", "B4", "B5", "B6")}
    if workbook_values != {"B2": 0, "B3": 0, "B4": 0, "B5": 0, "B6": 1}:
        raise RuntimeError(f"Exported values were not preserved: {workbook_values!r}")
    if workbook_formats["B2"] != "0.00" or workbook_formats["B4"] != "0.00":
        raise RuntimeError(f"Visible zero formats were not preserved: {workbook_formats!r}")
    if workbook_formats["B3"].split(";")[2] != "" or workbook_formats["B5"].split(";")[2] != "":
        raise RuntimeError(f"Hidden-zero formats were not preserved: {workbook_formats!r}")

    result = {
        "fixture": FIXTURE.name,
        "imported_grid_display": imported,
        "imported_page_display": page_values,
        "reopened_display": reopened_values,
        "exported_values": workbook_values,
        "exported_number_formats": workbook_formats,
        "output": str(output.relative_to(ROOT)),
    }
    result_path = OUT / "xls_zero_formula_display.json"
    result_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
